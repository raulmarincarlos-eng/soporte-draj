from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify
from werkzeug.security import check_password_hash
from functools import wraps
from database import init_db, get_db_connection
from datetime import datetime
import os
import sys
from werkzeug.utils import secure_filename
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flaskwebgui import FlaskUI

# Soporte para rutas de PyInstaller (sys._MEIPASS)
if getattr(sys, 'frozen', False):
    template_folder = os.path.join(sys._MEIPASS, 'templates')
    static_folder = os.path.join(sys._MEIPASS, 'static')
    app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
else:
    app = Flask(__name__)

UPLOAD_FOLDER = os.path.join('static', 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024 # 16 MB max
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf', 'doc', 'docx', 'xls', 'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


app.secret_key = 'Soporte_DRAJ_Seguridad_2026'

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Inicializar base de datos al arrancar
try:
    init_db()
except Exception as e:
    print(f"Error conectando a MongoDB: {e}")

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        db = get_db_connection()
        user = db.usuarios.find_one({"username": username})
        
        if user and check_password_hash(user['password'], password):
            session['user_id'] = str(user['_id'])
            session['username'] = user['username']
            session['nombre'] = user['nombre']
            return redirect(url_for('dashboard'))
        else:
            flash('Usuario o contraseña incorrectos.', 'danger')
            
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.context_processor
def inject_pendientes():
    if 'user_id' in session:
        try:
            db = get_db_connection()
            count = db.atenciones.count_documents({"resultado": "Pendiente"})
            return dict(total_pendientes=count)
        except:
            return dict(total_pendientes=0)
    return dict(total_pendientes=0)

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('solicitar'))

@app.route('/dashboard')
@login_required
def dashboard():
    db = get_db_connection()
    search = request.args.get('search', '')
    
    if search:
        query = {"$or": [
            {"area_usuaria": {"$regex": search, "$options": "i"}},
            {"id": {"$regex": search, "$options": "i"}}
        ]}
        # Compatibilidad con IDs numericos viejos
        if search.isdigit():
            query["$or"].append({"id": int(search)})
            query["$or"].append({"id_secuencial": int(search)})
            
        atenciones = list(db.atenciones.find(query).sort("id_secuencial", 1))
    else:
        atenciones = list(db.atenciones.find().sort("id_secuencial", 1))
    
    # Estadísticas
    today = datetime.now().strftime('%Y-%m-%d')
    total_hoy = db.atenciones.count_documents({"fecha_registro": today})
    
    current_month = datetime.now().strftime('%Y-%m')
    total_mes = db.atenciones.count_documents({"fecha_registro": {"$regex": f"^{current_month}"}})
    
    resultados_data = db.atenciones.aggregate([
        {"$group": {"_id": "$resultado", "count": {"$sum": 1}}}
    ])
    resultados_dict = {row["_id"]: row["count"] for row in resultados_data if row["_id"]}
    
    top_areas_data = list(db.atenciones.aggregate([
        {"$group": {"_id": "$area_usuaria", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 5}
    ]))
    for row in top_areas_data:
        row["area_usuaria"] = row["_id"]
        
    return render_template('dashboard.html', 
                           atenciones=atenciones, 
                           total_hoy=total_hoy, 
                           total_mes=total_mes, 
                           resultados_dict=resultados_dict, 
                           top_areas_data=top_areas_data,
                           search=search)

def generar_nuevo_id(db):
    last_atencion = db.atenciones.find_one(sort=[("id_secuencial", -1)])
    if not last_atencion:
        # Fallback para base de datos antigua
        last_atencion = db.atenciones.find_one(sort=[("id", -1)])
        
    if last_atencion:
        if "id_secuencial" in last_atencion:
            nuevo_num = last_atencion["id_secuencial"] + 1
        else:
            try:
                nuevo_num = int(last_atencion["id"]) + 1
            except:
                nuevo_num = 1
    else:
        nuevo_num = 1
        
    year = datetime.now().year
    # Formato: SOP-2026-0001
    codigo = f"SOP-{year}-{nuevo_num:04d}"
    return codigo, nuevo_num

@app.route('/nuevo', methods=('GET', 'POST'))
@login_required
def nuevo():
    db = get_db_connection()
    nuevo_codigo, nuevo_num = generar_nuevo_id(db)
    
    if request.method == 'POST':
        id_atencion = request.form.get('id_atencion')
        
        direccion = request.form.get('direccion', '')
        sub_area = request.form.get('sub_area', '')
        
        area_usuaria = sub_area if sub_area else direccion
        direccion_oficina = direccion
            
        file = request.files.get('evidencia_parcial')
        evidencia_filename = None
        
        if file and file.filename != '':
            if allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filename = f"{id_atencion}_{filename}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                evidencia_filename = filename

            
        doc = {
            "id": id_atencion, # Ej: SOP-2026-0001
            "id_secuencial": nuevo_num, # Para ordenar correctamente
            "fecha_registro": request.form.get('fecha_registro'),
            "hora_registro": request.form.get('hora_registro'),
            "area_usuaria": area_usuaria,
            "direccion_oficina": direccion_oficina,
            "tipo_atencion": request.form.get('tipo_atencion'),
            "descripcion": request.form.get('descripcion'),
            "clasif_software": bool(request.form.get('clasif_software')),
            "clasif_hardware": bool(request.form.get('clasif_hardware')),
            "clasif_red": bool(request.form.get('clasif_red')),
            "clasif_accesos": bool(request.form.get('clasif_accesos')),
            "clasif_biometrico": bool(request.form.get('clasif_biometrico')),
            "clasif_otros": bool(request.form.get('clasif_otros')),
            "fecha_inicio": request.form.get('fecha_inicio'),
            "fecha_fin": request.form.get('fecha_fin'),
            "detalle_actividades": request.form.get('detalle_actividades'),
            "resultado": request.form.get('resultado'),
            "observaciones": request.form.get('observaciones'),
            "conf_nombre": request.form.get('conf_nombre'),
            "conf_cargo": request.form.get('conf_cargo'),
            "conf_fecha": request.form.get('conf_fecha'),
            "resp_nombre": request.form.get('resp_nombre'),
            "resp_cargo": request.form.get('resp_cargo'),
            "resp_fecha": request.form.get('resp_fecha'),
            "resp_hora": request.form.get('resp_hora'),
            "ip_maquina": request.form.get('ip_maquina'),
            "evidencia_parcial": evidencia_filename
        }
        db.atenciones.insert_one(doc)
        flash('Atención registrada exitosamente', 'success')
        return redirect(url_for('dashboard'))
        
    return render_template('formulario.html', atencion=None, new_id=nuevo_codigo)

import socket

@app.route('/solicitar', methods=('GET', 'POST'))
def solicitar():
    if request.method == 'POST':
        db = get_db_connection()
        nuevo_codigo, nuevo_num = generar_nuevo_id(db)
        
        # Captura de IP y Navegador para validación legal
        ip_cliente = request.remote_addr
        user_agent = request.user_agent.string
        
        direccion = request.form.get('direccion', '')
        sub_area = request.form.get('sub_area', '')
        area_usuaria = sub_area if sub_area else direccion
        
        doc = {
            "id": nuevo_codigo,
            "id_secuencial": nuevo_num,
            "estado_firma": "Pendiente",
            "fecha_registro": datetime.now().strftime('%Y-%m-%d'),
            "hora_registro": datetime.now().strftime('%H:%M:%S'),
            "area_usuaria": area_usuaria,
            "direccion_oficina": direccion,
            "descripcion": request.form.get('descripcion'),
            "ip_cliente": ip_cliente,
            "user_agent": user_agent,
            "pc_cliente": "Registrado via Web",
            "contacto_usuario": request.form.get('contacto', ''),
            "equipo_usuario": request.form.get('equipo', ''),
            "tipo_atencion": request.form.get('tipo_atencion', ''),
            "clasif_software": bool(request.form.get('clasif_software')),
            "clasif_hardware": bool(request.form.get('clasif_hardware')),
            "clasif_red": bool(request.form.get('clasif_red')),
            "clasif_accesos": bool(request.form.get('clasif_accesos')),
            "clasif_biometrico": bool(request.form.get('clasif_biometrico')),
            "clasif_otros": bool(request.form.get('clasif_otros')),
            "fecha_inicio": "", "fecha_fin": "",
            "detalle_actividades": "", "resultado": "Pendiente", "observaciones": "Solicitud web en cola de espera.",
            "conf_nombre": request.form.get('nombre_solicitante', '').upper(),
            "conf_cargo": "", "conf_fecha": "", "resp_nombre": "", "resp_cargo": "", "resp_fecha": ""
        }
        db.atenciones.insert_one(doc)
        flash(f'Solicitud enviada. Tu N° de Ticket es: {nuevo_codigo}', 'success')
        return redirect(url_for('consultar', ticket_id=nuevo_codigo))
        
    return render_template('solicitar.html')

@app.route('/consultar', methods=['GET'])
def consultar():
    ticket_id = request.args.get('ticket_id', '').strip().upper()
    atencion = None
    error = None
    
    if ticket_id:
        db = get_db_connection()
        atencion = db.atenciones.find_one({"id": ticket_id})
        if not atencion:
            error = f"No se encontró ningún ticket con el código {ticket_id}. Verifica e intenta nuevamente."
            
    return render_template('consulta.html', atencion=atencion, search_id=ticket_id, error=error)

@app.route('/api/ticket/<ticket_id>', methods=['GET'])
def api_ticket(ticket_id):
    db = get_db_connection()
    atencion = db.atenciones.find_one({"id": ticket_id.upper().strip()})
    if atencion:
        return jsonify({
            "estado": atencion.get("resultado", "Pendiente"),
            "fecha": atencion.get("fecha_registro", ""),
            "area": atencion.get("area_usuaria", "")
        })
    return jsonify({"error": "No encontrado"}), 404

@app.route('/editar/<id_str>', methods=('GET', 'POST'))
@login_required
def editar(id_str):
    db = get_db_connection()
    # Buscar por string o por int (retrocompatibilidad)
    atencion = db.atenciones.find_one({"id": id_str})
    if not atencion and id_str.isdigit():
        atencion = db.atenciones.find_one({"id": int(id_str)})
        
    if not atencion:
        flash('Registro no encontrado', 'danger')
        return redirect(url_for('dashboard'))
    
    direccion_actual = atencion.get('direccion_oficina', atencion.get('area_usuaria', ''))
    sub_area_actual = ''
    if atencion.get('direccion_oficina') and atencion.get('area_usuaria') and atencion.get('area_usuaria') != atencion.get('direccion_oficina'):
        sub_area_actual = atencion.get('area_usuaria')
    
    if request.method == 'POST':
        direccion = request.form.get('direccion', '')
        sub_area = request.form.get('sub_area', '')
        
        area_usuaria = sub_area if sub_area else direccion
        direccion_oficina = direccion
            
        file = request.files.get('evidencia_parcial')
        evidencia_filename = None
        
        if file and file.filename != '':
            if allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filename = f"{id_atencion}_{filename}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                evidencia_filename = filename

        
        update_data = {
            "fecha_registro": request.form.get('fecha_registro'),
            "hora_registro": request.form.get('hora_registro'),
            "area_usuaria": area_usuaria,
            "direccion_oficina": direccion_oficina,
            "tipo_atencion": request.form.get('tipo_atencion'),
            "descripcion": request.form.get('descripcion'),
            "clasif_software": bool(request.form.get('clasif_software')),
            "clasif_hardware": bool(request.form.get('clasif_hardware')),
            "clasif_red": bool(request.form.get('clasif_red')),
            "clasif_accesos": bool(request.form.get('clasif_accesos')),
            "clasif_biometrico": bool(request.form.get('clasif_biometrico')),
            "clasif_otros": bool(request.form.get('clasif_otros')),
            "fecha_inicio": request.form.get('fecha_inicio'),
            "fecha_fin": request.form.get('fecha_fin'),
            "detalle_actividades": request.form.get('detalle_actividades'),
            "resultado": request.form.get('resultado'),
            "observaciones": request.form.get('observaciones'),
            "conf_nombre": request.form.get('conf_nombre'),
            "conf_cargo": request.form.get('conf_cargo'),
            "conf_fecha": request.form.get('conf_fecha'),
            "resp_nombre": request.form.get('resp_nombre'),
            "resp_cargo": request.form.get('resp_cargo'),
            "resp_fecha": request.form.get('resp_fecha'),
            "resp_hora": request.form.get('resp_hora'),
            "ip_maquina": request.form.get('ip_maquina'),
            "evidencia_parcial": evidencia_filename
        }
        db.atenciones.update_one({"_id": atencion["_id"]}, {"$set": update_data})
        flash('Registro actualizado correctamente.', 'success')
        return redirect(url_for('detalle', id_str=atencion["id"]))
        
    return render_template('formulario.html', atencion=atencion, direccion_actual=direccion_actual, sub_area_actual=sub_area_actual)

@app.route('/detalle/<id_str>')
@login_required
def detalle(id_str):
    db = get_db_connection()
    atencion = db.atenciones.find_one({"id": id_str})
    if not atencion and id_str.isdigit():
        atencion = db.atenciones.find_one({"id": int(id_str)})
        
    if atencion is None:
        flash('Registro no encontrado', 'danger')
        return redirect(url_for('dashboard'))
        
    return render_template('detalle.html', atencion=atencion)

@app.route('/exportar_excel')
@login_required
def exportar_excel():
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    
    db = get_db_connection()
    if fecha_inicio and fecha_fin:
        atenciones = list(db.atenciones.find({"fecha_registro": {"$gte": fecha_inicio, "$lte": fecha_fin}}).sort("id_secuencial", -1))
    else:
        atenciones = list(db.atenciones.find().sort("id_secuencial", -1))
        
    for a in atenciones:
        if '_id' in a: del a['_id']
        if 'id_secuencial' in a: del a['id_secuencial']
        if 'ip_cliente' in a: del a['ip_cliente']
        if 'user_agent' in a: del a['user_agent']
            
    if not atenciones:
        flash('No hay registros en el rango seleccionado.', 'warning')
        return redirect(url_for('dashboard'))
        
    df = pd.DataFrame(atenciones)
    
    nombres_columnas = {
        'id': 'Código',
        'fecha_registro': 'Fecha Registro',
        'hora_registro': 'Hora',
        'area_usuaria': 'Área Usuaria',
        'direccion_oficina': 'Oficina/Dirección',
        'tipo_atencion': 'Tipo de Atención',
        'descripcion': 'Descripción del Requerimiento',
        'clasif_software': 'Software',
        'clasif_hardware': 'Hardware',
        'clasif_red': 'Red/Internet',
        'clasif_accesos': 'Accesos',
        'clasif_biometrico': 'Biométrico',
        'clasif_otros': 'Otros (Clasif)',
        'fecha_inicio': 'Fecha Inicio Atención',
        'fecha_fin': 'Fecha Fin Atención',
        'detalle_actividades': 'Detalle de Actividades',
        'resultado': 'Resultado',
        'observaciones': 'Observaciones',
        'conf_nombre': 'Usuario Conformidad',
        'conf_cargo': 'Cargo Usuario',
        'conf_fecha': 'Fecha Conformidad',
        'resp_nombre': 'Técnico Responsable',
        'resp_cargo': 'Cargo Técnico',
        'resp_fecha': 'Fecha Cierre',
        'resp_hora': 'Hora Cierre',
        'ip_maquina': 'IP Máquina'
    }
    # Filtrar solo las columnas que queremos en el Excel
    columnas_deseadas = [k for k in nombres_columnas.keys() if k in df.columns]
    df = df[columnas_deseadas]
    
    nombres_columnas_existentes = {k: v for k in columnas_deseadas for k, v in nombres_columnas.items() if k in df.columns}
    df.rename(columns=nombres_columnas, inplace=True)
    
    for col in ['Software', 'Hardware', 'Red/Internet', 'Accesos', 'Biométrico', 'Otros (Clasif)']:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: 'SI' if x == 1 or x == '1' or x == True else 'NO')
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Reporte DRAJ')
        
        workbook = writer.book
        worksheet = writer.sheets['Reporte DRAJ']
        
        header_fill = PatternFill(start_color="0A58CA", end_color="0A58CA", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
        for col in worksheet.columns:
            max_length = 0
            column = [cell for cell in col]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            adjusted_width = (max_length + 2)
            if adjusted_width > 50: adjusted_width = 50
            if adjusted_width < 12: adjusted_width = 12
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    output.seek(0)
    
    if fecha_inicio and fecha_fin:
        nombre_archivo = f'Reporte_Soporte_{fecha_inicio}_al_{fecha_fin}.xlsx'
    else:
        nombre_archivo = f'Reporte_Soporte_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
    return send_file(output, download_name=nombre_archivo, as_attachment=True)

@app.route('/eliminar/<id_str>', methods=['POST'])
@login_required
def eliminar(id_str):
    db = get_db_connection()
    # Buscar por string o por int (retrocompatibilidad)
    atencion = db.atenciones.find_one({"id": id_str})
    if not atencion and id_str.isdigit():
        atencion = db.atenciones.find_one({"id": int(id_str)})
        
    if atencion:
        db.atenciones.delete_one({"_id": atencion["_id"]})
        flash(f'Atención {id_str} eliminada correctamente.', 'success')
    return redirect(url_for('dashboard'))

from xhtml2pdf import pisa
import io

@app.route('/exportar_pdf/<id_str>')
@login_required
def exportar_pdf(id_str):
    db = get_db_connection()
    atencion = db.atenciones.find_one({"id": id_str})
    if not atencion and id_str.isdigit():
        atencion = db.atenciones.find_one({"id": int(id_str)})
        
    if not atencion:
        flash('Registro no encontrado', 'danger')
        return redirect(url_for('dashboard'))
        
    html = render_template('pdf_conformidad.html', atencion=atencion)
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        result.seek(0)
        return send_file(result, download_name=f'Conformidad_{atencion["id"]}.pdf', as_attachment=True, mimetype='application/pdf')
    else:
        flash('Error al generar el PDF', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/voucher_pdf/<id_str>')
def voucher_pdf(id_str):
    db = get_db_connection()
    atencion = db.atenciones.find_one({"id": id_str})
    if not atencion and id_str.isdigit():
        atencion = db.atenciones.find_one({"id": int(id_str)})
        
    if not atencion:
        flash('Ticket no encontrado', 'danger')
        return redirect(url_for('solicitar'))
        
    html = render_template('pdf_voucher.html', atencion=atencion)
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        result.seek(0)
        return send_file(result, download_name=f'Ticket_Soporte_{atencion["id"]}.pdf', as_attachment=True, mimetype='application/pdf')
    else:
        flash('Error al generar el PDF del Voucher', 'danger')
        return redirect(url_for('consultar', ticket_id=id_str))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
